"""
Chat with Data — Python micro-service (Standard + Variance)
-----------------------------------------------------------
Repackages the prototype's proven analytical core as a stateless-style REST API
that the .NET front end calls over the internal network.

The analytical core is UNCHANGED from the prototype:
  - services/llm.py        two-call LLM pipeline (Call 1 code, Call 2 answer)
  - services/executor.py   executes generated pandas code in one namespace
  - services/file_handler.py CSV/Excel reading, schema extraction
  - prompts/standard.py, prompts/variance.py   the exact mode prompts

The only architectural change vs. the prototype's routes/chat.py is that the
session id arrives in the `X-Session-Id` header (managed by the .NET layer)
instead of a Flask cookie. The pipeline itself is reproduced verbatim.

For local validation, session state is held in memory (services/data_manager.py).
In production this service runs under gunicorn with SQL Server + a file share
providing durable, stateless session/data storage (see README / architecture doc).
"""

import io
import json
import os
import uuid
from datetime import datetime, timedelta
from typing import Any

import pandas as pd
import requests as _requests
from flask import Blueprint, Flask, jsonify, request, send_file

from config import Config
from logger import logger
from models.schemas import DebugEntry
from prompts import standard, variance, linking
from services import llm as llm_service
from services import insights
from services import request_context, usage_capture
from services import database as db
from services.admin_auth import is_admin, require_admin
from services.crypto import CryptoNotConfigured, decrypt_api_key, encrypt_api_key, mask_api_key
from services.data_manager import get_session, save_session, get_schema_info
from services.executor import execute_generated_code
from services.file_handler import FileHandlerError, get_excel_sheets, read_file
from services.model_manager import (
    change_model, get_available_models, get_current_model_config, get_model_history,
)
from services import experts as experts_svc
from services import data_sources
from services.report_export import build_conversation_docx

api = Blueprint("api", __name__, url_prefix="/api")


# ── Session id (provided by the .NET front end) ───────────────────────────────

class MissingSessionId(Exception):
    pass


def _sid() -> str:
    """
    The .NET presentation layer owns the user session and passes its id in the
    X-Session-Id header. Fall back to a query/form value for direct testing.
    """
    sid = (
        request.headers.get("X-Session-Id")
        or request.values.get("session_id")
    )
    if not sid:
        raise MissingSessionId()
    return sid


def _username() -> str:
    """
    Identity supplied by the .NET layer in X-User: the Windows username when
    Windows Authentication is on, otherwise the .NET host's dev fallback.
    """
    return (request.headers.get("X-User") or "unknown").strip()


def _groups() -> list:
    """
    The caller's AD groups, resolved by .NET from the Windows sign-in badge
    (only groups that Data Experts actually reference are sent). Empty when
    auth is off, unless simulated via the .NET Auth:DevGroups setting.
    """
    raw = request.headers.get("X-User-Groups") or ""
    return [g.strip() for g in raw.replace(";", ",").split(",") if g.strip()]


@api.before_request
def _check_internal_secret():
    """
    Optional shared secret between .NET and Python (INTERNAL_API_SECRET in
    .env). Off by default. When set, every /api request must carry the matching
    X-Internal-Auth header — closes the 'anyone on the network can hit port
    8000 and claim any X-User' gap while Python runs on a separate machine.
    """
    secret = Config.INTERNAL_API_SECRET
    if secret and request.headers.get("X-Internal-Auth") != secret:
        return jsonify({"error": "Unauthorized"}), 401


def _resolve_user_context(username: str, session=None):
    """
    Look up the user, decrypt their API key, and read the active model.
    Checks session for user's model preference, falls back to admin default.
    Returns ((user, api_key, model_name), None) on success or (None, flask_response).
    """
    user = db.get_user_by_username(username)
    if not user or not user.get("IsActive", 1):
        return None, (jsonify({
            "error": "No API key on file. Please set up your LiteLLM API key to use the app.",
            "needs_key": True,
        }), 401)
    try:
        api_key = decrypt_api_key(user["APIKey"])
    except CryptoNotConfigured as e:
        return None, (jsonify({"error": str(e)}), 503)
    except Exception:
        logger.error(f"Could not decrypt stored key for {username}")
        return None, (jsonify({
            "error": "Your stored API key could not be read. Please re-enter it.",
            "needs_key": True,
        }), 401)
    # Check session for user's model preference, fall back to admin default
    admin_default = get_current_model_config()["model_name"]
    model_name = admin_default
    if session and session.user_model_preference:
        model_name = session.user_model_preference
        logger.info(f"Using user-selected model: {model_name} (admin default: {admin_default})")
    return (user, api_key, model_name), None


# ── Shared pipeline helpers (reproduced verbatim from prototype chat.py) ───────

def _get_prev_result_meta(raw_result: Any, max_chars: int = 1000) -> str:
    """Truncated string representation of previous result for LLM context."""
    if raw_result is None:
        return None
    try:
        if isinstance(raw_result, pd.DataFrame):
            if raw_result.empty:
                return "Empty DataFrame (0 rows)"
            preview = raw_result.head(5).to_string(index=False)
            result_str = f"DataFrame with {len(raw_result)} rows, {len(raw_result.columns)} columns:\n{preview}"
            if len(raw_result) > 5:
                result_str += f"\n... ({len(raw_result) - 5} more rows)"
        elif isinstance(raw_result, pd.Series):
            preview = raw_result.head(5).to_string()
            result_str = f"Series with {len(raw_result)} values:\n{preview}"
            if len(raw_result) > 5:
                result_str += f"\n... ({len(raw_result) - 5} more values)"
        elif isinstance(raw_result, (list, tuple)):
            preview = "\n".join(str(item) for item in raw_result[:5])
            result_str = f"List with {len(raw_result)} items:\n{preview}"
            if len(raw_result) > 5:
                result_str += f"\n... ({len(raw_result) - 5} more items)"
        elif isinstance(raw_result, dict):
            keys_info = ", ".join(f"{k}: {type(v).__name__}" for k, v in list(raw_result.items())[:5])
            result_str = f"Dict with {len(raw_result)} keys: {keys_info}"
        else:
            result_str = str(raw_result)

        if len(result_str) > max_chars:
            result_str = result_str[:max_chars] + "... [truncated]"
        return result_str
    except Exception as e:
        logger.warning(f"Error building prev_result_meta: {e}")
        return None


def _inject_error(messages: list, error: str) -> list:
    updated  = messages.copy()
    last_msg = updated[-1].copy()
    last_msg["content"] += f"\n\nYour previous code failed:\n{error}\n\nReturn ONLY fixed Python code. No prose."
    updated[-1] = last_msg
    return updated


def _run_pipeline(code_messages: list, answer_messages_fn, debug: list,
                  prev_result: Any = None, **dataframes) -> tuple:
    """
    Shared two-call pipeline with one auto-retry (logic verbatim from prototype).
    Also returns the computed result_str/metadata so the "one sentence / more
    detail" refine feature can re-run Call 2 against the SAME computed result.
    Returns (success, answer, chart, debug, raw_result, result_str, metadata).
    """
    # ── Call 1: generate code ──────────────────────────────────────────────────
    ok, code = llm_service.generate_query_code(code_messages)
    if not ok:
        return False, code, None, debug, None, None, None, None
    debug.append(DebugEntry("Generated code", code))

    # ── Execute ────────────────────────────────────────────────────────────────
    success, result_str, raw_result, metadata = execute_generated_code(code, prev_result=prev_result, **dataframes)

    if not success:
        debug.append(DebugEntry("Execution error — retrying", result_str))
        retry_messages = _inject_error(code_messages, result_str)
        ok2, code2 = llm_service.generate_query_code(retry_messages)
        if ok2:
            debug.append(DebugEntry("Retry code", code2))
            success, result_str, raw_result, metadata = execute_generated_code(code2, prev_result=prev_result, **dataframes)
        if not success:
            debug.append(DebugEntry("Retry also failed", result_str))
            return False, "I wasn't able to compute an answer for that question. Try rephrasing it.", None, debug, None, None, None, None

    debug.append(DebugEntry("Query result", result_str))

    # ── Call 2: generate answer ────────────────────────────────────────────────
    answer_messages = answer_messages_fn(result_str, metadata)
    ok, raw_answer  = llm_service.generate_human_answer(answer_messages)
    if not ok:
        return False, raw_answer, None, debug, None, None, None, None

    answer, chart, formatting = llm_service.parse_answer_response(raw_answer)
    debug.append(DebugEntry("Final answer", answer))
    return True, answer, chart, debug, raw_result, result_str, metadata, formatting


# ── Upload (Standard + Variance) ──────────────────────────────────────────────

def _analyze_join_suggestions(sess, new_slot: str):
    """
    Analyze potential join keys between existing files and newly uploaded file.
    Returns list of suggested joins sorted by confidence (0-1 score).
    """
    from difflib import SequenceMatcher

    def get_column_priority(col_name: str) -> int:
        """
        Assign priority score to column names.
        Higher scores = better join key candidates.

        Priority levels:
        - 3 (HIGH): ID fields, codes, keys
        - 2 (MEDIUM): Numbers, identifiers
        - 1 (LOW): Descriptive fields (name, description)
        - 0 (VERY LOW): Aggregated values (cost, amount, total, date)
        """
        col_lower = str(col_name).lower()

        # HIGH priority: ID fields and codes
        if col_lower.endswith('id') or col_lower.endswith('_id'):
            return 3
        if col_lower.endswith('code') or col_lower.endswith('_code'):
            return 3
        if 'key' in col_lower or col_lower.endswith('num') or col_lower.endswith('number'):
            return 3
        if col_lower in ['id', 'code', 'key', 'identifier']:
            return 3

        # VERY LOW priority: Aggregated values and dates (bad join keys)
        if any(x in col_lower for x in ['cost', 'amount', 'total', 'price', 'value', 'sum']):
            return 0
        if any(x in col_lower for x in ['date', 'time', 'period', 'year', 'month', 'day']):
            return 0

        # LOW priority: Descriptive text fields
        if any(x in col_lower for x in ['name', 'description', 'desc', 'title', 'label']):
            return 1

        # MEDIUM priority: Everything else (category fields, etc.)
        return 2

    suggestions = []
    new_df = getattr(sess, f"df_{new_slot}", None)
    new_label = getattr(sess, f"label_{new_slot}", f"File {new_slot.upper()}")

    if new_df is None:
        return suggestions

    new_cols = set(new_df.columns)

    # Value overlap is computed on a SAMPLE of distinct values per column —
    # plenty to detect a real link, and instant even on 500k-row files
    # (previously every value of every column pair was compared).
    SAMPLE = 5000
    _vals_cache = {}

    def sample_values(slot, df, col):
        """Up to SAMPLE distinct values of a column, as strings (cached per column)."""
        key = (slot, col)
        if key not in _vals_cache:
            vals = df[col].dropna().unique()[:SAMPLE]
            _vals_cache[key] = set(map(str, vals))
        return _vals_cache[key]

    _str_cache = {}

    def str_column(slot, df, col):
        """Full column as a string Series (cached) for vectorized membership tests."""
        key = (slot, col)
        if key not in _str_cache:
            _str_cache[key] = df[col].dropna().astype(str)
        return _str_cache[key]

    def value_overlap(slot_a, df_a_, col_a, slot_b, df_b_, col_b):
        """
        Estimated share of col_a's sampled distinct values that appear anywhere
        in col_b. Sampling one side but scanning the full other side keeps this
        instant on large files AND correct when files are sorted differently.
        """
        sample = sample_values(slot_a, df_a_, col_a)
        if not sample:
            return 0.0
        other = str_column(slot_b, df_b_, col_b)
        if other.empty:
            return 0.0
        hits = other[other.isin(sample)].nunique()
        return hits / len(sample)

    # Compare with all other uploaded files
    for other_slot in ["a", "b", "c", "d"]:
        if other_slot == new_slot:
            continue

        other_df = getattr(sess, f"df_{other_slot}", None)
        if other_df is None:
            continue

        other_label = getattr(sess, f"label_{other_slot}", f"File {other_slot.upper()}")
        other_cols = set(other_df.columns)

        # Find column matches (exact or fuzzy)
        for new_col in new_cols:
            for other_col in other_cols:
                # Exact match
                if new_col == other_col:
                    # Bad join-key types (costs, amounts, dates) are never
                    # suggested — skip the value scan entirely.
                    if get_column_priority(new_col) == 0 or get_column_priority(other_col) == 0:
                        continue
                    # Estimate value overlap: sampled distinct values of the new
                    # column, membership-tested against the full other column.
                    overlap = value_overlap(new_slot, new_df, new_col, other_slot, other_df, other_col)
                    if overlap > 0.1:  # At least 10% overlap
                        priority = (get_column_priority(new_col) + get_column_priority(other_col)) / 2
                        suggestions.append({
                            "file1": new_label,
                            "file2": other_label,
                            "slot1": new_slot,
                            "slot2": other_slot,
                            "column1": new_col,
                            "column2": other_col,
                            "confidence": overlap,
                            "match_type": "exact",
                            "priority": priority,
                        })
                # Fuzzy match (only for different column names)
                elif new_col != other_col:
                    # Bad join-key types are never suggested — skip early.
                    if get_column_priority(new_col) == 0 or get_column_priority(other_col) == 0:
                        continue
                    # Apply abbreviation expansion before fuzzy matching
                    COMMON_ABBREV = {
                        'num': 'number', 'acct': 'account', 'cust': 'customer',
                        'qty': 'quantity', 'amt': 'amount', 'dt': 'date',
                        'id': 'identifier', 'desc': 'description', 'addr': 'address',
                        'proj': 'project', 'emp': 'employee', 'dept': 'department'
                    }

                    def expand_abbrev(col_name):
                        """Expand common abbreviations in column names."""
                        col_lower = str(col_name).lower()
                        for abbrev, full in COMMON_ABBREV.items():
                            col_lower = col_lower.replace(abbrev, full)
                        return col_lower

                    new_col_expanded = expand_abbrev(new_col)
                    other_col_expanded = expand_abbrev(other_col)

                    similarity = SequenceMatcher(None, new_col_expanded, other_col_expanded).ratio()
                    if similarity > 0.6:  # 60% similarity threshold (lowered from 70%)
                        # Estimate value overlap (sampled, vectorized, sort-safe)
                        overlap = value_overlap(new_slot, new_df, new_col, other_slot, other_df, other_col)
                        if overlap > 0.1:
                            priority = (get_column_priority(new_col) + get_column_priority(other_col)) / 2
                            suggestions.append({
                                "file1": new_label,
                                "file2": other_label,
                                "slot1": new_slot,
                                "slot2": other_slot,
                                "column1": new_col,
                                "column2": other_col,
                                "confidence": overlap * similarity,  # Combined score
                                "match_type": "fuzzy",
                                "priority": priority,
                            })

    # Remove true duplicates (same files + same columns)
    # Keep suggestions that show different file pairings
    seen_pairs = set()
    unique_suggestions = []
    for sug in suggestions:
        # Key includes both column names AND file slots to preserve different pairings
        pair_key = (
            tuple(sorted([sug["slot1"], sug["slot2"]])),
            tuple(sorted([sug["column1"], sug["column2"]]))
        )
        if pair_key not in seen_pairs:
            seen_pairs.add(pair_key)
            unique_suggestions.append(sug)

    # Sort by multiple criteria (all descending):
    # 1. Match type (exact first)
    # 2. Column priority (ID/Code fields first)
    # 3. Confidence (overlap score)
    unique_suggestions.sort(
        key=lambda x: (x["match_type"] == "exact", x["priority"], x["confidence"]),
        reverse=True
    )
    return unique_suggestions[:6]


def _handle_upload(file, slot: str, sess, sheet_name: str = None) -> dict:
    """Shared upload logic (verbatim from prototype upload.py)."""
    if not file or not file.filename:
        return {"error": "No file provided"}

    file_bytes = file.read()

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext in ("xlsx", "xls") and not sheet_name:
        sheets = get_excel_sheets(file_bytes)
        if len(sheets) > 1:
            return {"needs_sheet_selection": True, "sheets": sheets,
                    "filename": file.filename, "slot": slot}

    try:
        df, encoding, warnings = read_file(file_bytes, file.filename, sheet_name)
    except FileHandlerError as e:
        return {"error": str(e)}

    schema = get_schema_info(df)

    setattr(sess, slot if slot == "df" else f"df_{slot}", df)
    setattr(sess, "schema" if slot == "df" else f"schema_{slot}", schema)
    if slot == "df":
        sess.filename = file.filename

    columns = [{"name": str(col), "type": str(df[col].dtype)} for col in df.columns]
    preview = df.head(5).fillna("").astype(str).replace("nan", "").to_dict(orient="records")

    logger.info(f"File uploaded | slot={slot} | file={file.filename} | shape={df.shape} | encoding={encoding}")

    # Upload-time insights — pure pandas, no LLM calls (services/insights.py).
    mode = "standard" if slot == "df" else "variance"
    try:
        profile     = insights.build_profile(df)
        anomalies   = insights.find_anomalies(df)
        suggestions = insights.suggest_questions(
            df, mode=mode,
            label_a=getattr(sess, "label_a", None),
            label_b=getattr(sess, "label_b", None),
        )
    except Exception as e:
        logger.warning(f"Insights failed (non-fatal): {e}")
        profile, anomalies, suggestions = None, [], []

    return {
        "ok": True, "rows": len(df), "cols": len(df.columns),
        "columns": columns, "preview": preview, "filename": file.filename,
        "encoding": encoding, "warnings": warnings,
        "profile": profile, "anomalies": anomalies, "suggestions": suggestions,
    }


@api.route("/upload/standard", methods=["POST"])
def upload_standard():
    sid  = _sid()
    sess = get_session(sid)
    # Tab-scoped clear: a new Standard file only resets the Standard conversation.
    sess.clear_history("standard")
    result = _handle_upload(request.files.get("file"), "df", sess, request.form.get("sheet_name"))
    if "ok" in result:
        save_session(sid, sess)
    return jsonify(result), (400 if "error" in result else 200)


@api.route("/upload/variance/<slot>", methods=["POST"])
def upload_variance(slot: str):
    if slot not in ("a", "b", "c", "d"):
        return jsonify({"error": "Invalid slot. Use 'a', 'b', 'c', or 'd'"}), 400
    sid  = _sid()
    sess = get_session(sid)
    # Tab-scoped clear: a new Variance file only resets the Variance conversation.
    sess.clear_history("variance")
    label = request.form.get("label", f"File {slot.upper()}")
    setattr(sess, f"label_{slot}", label)
    result = _handle_upload(request.files.get("file"), slot, sess, request.form.get("sheet_name"))
    if "ok" in result:
        # Links are file-bound: a new file in this slot invalidates any
        # accepted joins that referenced the previous file's columns.
        stale = [h for h in sess.join_hints if slot in (h.get("slot1"), h.get("slot2"))]
        if stale:
            sess.join_hints = [h for h in sess.join_hints if h not in stale]
            logger.info(f"Cleared {len(stale)} join link(s) referencing replaced slot '{slot}'")
        # Analyze join suggestions for multi-file linking (for any slot after the first)
        # This triggers when uploading File 2, 3, or 4
        if slot in ("b", "c", "d"):
            join_suggestions = _analyze_join_suggestions(sess, slot)
            result["join_suggestions"] = join_suggestions
        save_session(sid, sess)
    return jsonify(result), (400 if "error" in result else 200)


@api.route("/variance/accept_join", methods=["POST"])
def accept_join():
    """User accepted a suggested join - store it in join_hints for the prompt."""
    sid = _sid()
    sess = get_session(sid)
    body = request.get_json() or {}

    join_hint = {
        "slot1": body.get("slot1"),
        "slot2": body.get("slot2"),
        "column1": body.get("column1"),
        "column2": body.get("column2"),
        "file1": body.get("file1"),
        "file2": body.get("file2"),
    }

    # Add to join hints if not already present
    if join_hint not in sess.join_hints:
        sess.join_hints.append(join_hint)

    save_session(sid, sess)
    logger.info(f"Join accepted | {join_hint['file1']}.{join_hint['column1']} ↔ {join_hint['file2']}.{join_hint['column2']}")

    return jsonify({"ok": True, "message": f"Link saved: {join_hint['file1']}.{join_hint['column1']} ↔ {join_hint['file2']}.{join_hint['column2']}"}), 200


@api.route("/variance/remove_join", methods=["POST"])
def remove_join():
    """Remove one accepted join hint — the ✕ on an active link in the UI."""
    sid = _sid()
    sess = get_session(sid)
    body = request.get_json() or {}
    slot1, slot2 = body.get("slot1"), body.get("slot2")
    col1 = (body.get("column1") or "").lower()
    col2 = (body.get("column2") or "").lower()

    def is_match(h):
        a = (h.get("slot1"), (h.get("column1") or "").lower(),
             h.get("slot2"), (h.get("column2") or "").lower())
        return a == (slot1, col1, slot2, col2) or a == (slot2, col2, slot1, col1)

    before = len(sess.join_hints)
    sess.join_hints = [h for h in sess.join_hints if not is_match(h)]
    save_session(sid, sess)
    removed = before - len(sess.join_hints)
    logger.info(f"Join removed | sid={sid} | removed={removed}")
    return jsonify({"ok": True, "removed": removed, "joins": sess.join_hints})


@api.route("/variance/set_mode", methods=["POST"])
def set_manual_mode():
    """Allow user to manually select variance or linking mode."""
    sid = _sid()
    sess = get_session(sid)
    body = request.get_json() or {}

    mode = body.get("mode")  # 'variance', 'linking', or null for auto
    if mode not in (None, "variance", "linking"):
        return jsonify({"error": "Invalid mode. Use 'variance', 'linking', or null"}), 400

    sess.manual_mode = mode
    save_session(sid, sess)
    logger.info(f"Manual mode set to: {mode}")

    return jsonify({"ok": True, "mode": mode}), 200


@api.route("/variance/get_joins", methods=["GET"])
def get_active_joins():
    """Return the list of active joins for display in the UI."""
    sid = _sid()
    sess = get_session(sid)
    return jsonify({"joins": sess.join_hints}), 200


# ── Ask (the two-call pipeline) ───────────────────────────────────────────────

@api.route("/ask", methods=["POST"])
def ask():
    sid  = _sid()
    sess = get_session(sid)

    body     = request.get_json() or {}
    question = body.get("question", "").strip()
    mode     = body.get("mode", "standard")
    if not question:
        return jsonify({"error": "Empty question"}), 400

    # ── Per-user key + model + usage tracking (additive envelope) ─────────────
    # The pipeline below is UNCHANGED; this wrapper only sets the per-request
    # context that config.py serves to the untouched llm.py, and flushes the
    # captured token usage to the database afterwards.
    username = _username()
    ctx, err = _resolve_user_context(username, sess)
    if err:
        return err
    user, user_key, model_name = ctx
    request_id = uuid.uuid4().hex
    ctx_tokens = request_context.begin(user_key, model_name)
    try:
        return _ask_pipeline(sid, sess, question, mode)
    finally:
        usage_capture.flush(user["UserID"], username, sid, request_id, mode, question)
        request_context.end(ctx_tokens)


def _parse_simple_join(question: str, sess) -> list:
    """
    Extract join hints from EXPLICIT linking language in the question.

    Patterns supported (must use the word "link" or "join"):
    - "Link File 1 CustomerID to File 2 CustID"
    - "Link CustomerID to CustID" / "Join on AccountNum to AcctID"

    Every parsed column is validated case-insensitively against the real
    columns of the loaded files and emitted with its true casing; hints
    naming columns that don't exist are discarded. Plain "A = B" text is
    deliberately NOT treated as a link — users use equals signs to filter.

    Returns list of join hint dicts compatible with linking.py
    """
    import re

    def real_column(slot: str, name: str):
        """Case-insensitive lookup in the slot's DataFrame; returns the true column name or None."""
        df = getattr(sess, f"df_{slot}", None)
        if df is None:
            return None
        return {str(c).lower(): str(c) for c in df.columns}.get(name.lower())

    def label(slot: str) -> str:
        return getattr(sess, f"label_{slot}", f"File {slot.upper()}")

    uploaded = [s for s in ("a", "b", "c", "d") if getattr(sess, f"df_{s}", None) is not None]
    slot_map = {"1": "a", "2": "b", "3": "c", "4": "d"}
    hints, seen = [], set()

    def add_hint(slot1: str, slot2: str, col1: str, col2: str):
        key = (slot1, slot2, col1.lower(), col2.lower())
        if key in seen:
            return
        seen.add(key)
        hints.append({
            "slot1": slot1, "slot2": slot2,
            "column1": col1, "column2": col2,
            "file1": label(slot1), "file2": label(slot2),
        })

    def find_pair(col1: str, col2: str):
        """First pair of distinct loaded files where col1 and col2 actually exist."""
        for s1 in uploaded:
            for s2 in uploaded:
                if s1 == s2:
                    continue
                r1, r2 = real_column(s1, col1), real_column(s2, col2)
                if r1 and r2:
                    return s1, s2, r1, r2
        return None

    # Pattern 1: "link/join file X ColumnA to file Y ColumnB"
    pattern1 = r'(?:link|join)\s+file\s+([1-4])\s+(\w+)\s+(?:to|on|with)\s+file\s+([1-4])\s+(\w+)'
    explicit = re.findall(pattern1, question, re.IGNORECASE)
    for file1_num, col1, file2_num, col2 in explicit:
        slot1, slot2 = slot_map[file1_num], slot_map[file2_num]
        real1, real2 = real_column(slot1, col1), real_column(slot2, col2)
        if slot1 != slot2 and real1 and real2:
            add_hint(slot1, slot2, real1, real2)

    # Pattern 2: "link/join ColumnA to ColumnB" (no file numbers) — find the
    # pair of files where both columns actually exist. Skipped when pattern 1
    # matched so the same phrase isn't parsed twice.
    if not explicit:
        pattern2 = r'(?:link|join)\s+(?:on\s+)?(\w+)\s+(?:to|on|with)\s+(\w+)'
        for col1, col2 in re.findall(pattern2, question, re.IGNORECASE):
            if col1.lower() == "file":   # fragment of a malformed pattern-1 phrase
                continue
            found = find_pair(col1, col2)
            if found:
                add_hint(*found)

    if hints:
        logger.info(f"Parsed {len(hints)} validated join hints from question: {hints}")
    return hints


def _detect_analysis_mode(sess) -> str:
    """
    Determine if this is traditional variance or multi-file linking.

    Returns "variance" if:
    - Only files A & B uploaded
    - Schemas are 80%+ similar
    - No explicit join hints

    Returns "linking" if:
    - 3+ files uploaded
    - OR 2 files with different schemas
    - OR explicit join hints present
    - OR user manually selected linking mode
    """
    # Check for manual mode override
    if sess.manual_mode:
        logger.info(f"Mode detection: {sess.manual_mode} (manual override)")
        return sess.manual_mode

    # Count uploaded files
    uploaded_slots = []
    for slot in ["a", "b", "c", "d"]:
        if getattr(sess, f"df_{slot}", None) is not None:
            uploaded_slots.append(slot)

    # If 3+ files, definitely linking mode
    if len(uploaded_slots) >= 3:
        logger.info("Mode detection: linking (3+ files)")
        return "linking"

    # If explicit join hints exist, use linking mode
    if sess.join_hints:
        logger.info("Mode detection: linking (join hints present)")
        return "linking"

    # If only 2 files (a and b), check schema similarity
    if len(uploaded_slots) == 2 and "a" in uploaded_slots and "b" in uploaded_slots:
        schema_a = set(sess.df_a.columns)
        schema_b = set(sess.df_b.columns)
        common_cols = schema_a & schema_b
        total_cols = schema_a | schema_b
        similarity = len(common_cols) / len(total_cols) if total_cols else 0

        if similarity >= 0.8:
            logger.info(f"Mode detection: variance (schema similarity {similarity:.0%})")
            return "variance"
        else:
            logger.info(f"Mode detection: linking (schema similarity {similarity:.0%} < 80%)")
            return "linking"

    # Two files that are NOT in slots A+B (e.g., slots 1 & 3): the classic
    # variance path is hard-wired to df_a/df_b, so route any other pair to
    # linking mode, which gathers whichever slots are loaded.
    if len(uploaded_slots) == 2:
        logger.info(f"Mode detection: linking (2 files in slots {uploaded_slots}, not A+B)")
        return "linking"

    logger.info("Mode detection: variance (default)")
    return "variance"


def _ask_pipeline(sid, sess, question, mode):
    """The original /ask body, moved verbatim (logic identical to the prototype flow)."""
    debug = []
    logger.info(f"Question received | mode={mode} | sid={sid} | q={question[:80]}")
    history = sess.get_history(mode)

    if mode == "standard":
        if sess.df is None:
            return jsonify({"error": "No data loaded. Please upload a file first."}), 400
        prev_result = getattr(sess, "last_result", None)
        prev_result_meta = _get_prev_result_meta(prev_result, max_chars=1000)
        code_msgs = standard.build_code_gen_prompt(
            sess.schema, question, error_feedback=None,
            history=history, prev_result_meta=prev_result_meta,
        )
        success, answer, chart, debug, raw_result, result_str, metadata, formatting = _run_pipeline(
            code_msgs,
            lambda r, m: standard.build_answer_gen_prompt(question, r, history, m),
            debug, prev_result=prev_result, df=sess.df,
        )

    elif mode == "variance":
        # Check minimum files uploaded (at least 2)
        uploaded_count = sum(1 for slot in ["a", "b", "c", "d"] if getattr(sess, f"df_{slot}", None) is not None)
        if uploaded_count < 2:
            return jsonify({"error": "Please upload at least 2 files before asking questions."}), 400

        # Detect analysis mode: variance (same schema comparison) vs linking (multi-file join)
        analysis_mode = _detect_analysis_mode(sess)
        prev_result = getattr(sess, "last_result", None)
        prev_result_meta = _get_prev_result_meta(prev_result, max_chars=1000)

        if analysis_mode == "variance":
            # EXISTING CODE - traditional 2-file variance analysis
            code_msgs = variance.build_code_gen_prompt(
                sess.schema_a, sess.schema_b, sess.label_a, sess.label_b, question,
                error_feedback=None, history=history, prev_result_meta=prev_result_meta,
            )
            success, answer, chart, debug, raw_result, result_str, metadata, formatting = _run_pipeline(
                code_msgs,
                lambda r, m: variance.build_answer_gen_prompt(question, r, sess.label_a, sess.label_b, history, m),
                debug, prev_result=prev_result, df_a=sess.df_a, df_b=sess.df_b,
            )

        elif analysis_mode == "linking":
            # NEW CODE - multi-file linking analysis
            # Gather all uploaded dataframes, schemas, and labels
            dfs = {}
            schemas = {}
            labels = {}
            for slot in ["a", "b", "c", "d"]:
                df = getattr(sess, f"df_{slot}", None)
                if df is not None:
                    dfs[slot] = df
                    schemas[slot] = getattr(sess, f"schema_{slot}")
                    labels[slot] = getattr(sess, f"label_{slot}")

            # Parse natural language join hints from question
            parsed_hints = _parse_simple_join(question, sess)

            # Merge with stored join hints (accepted from UI)
            all_hints = sess.join_hints + parsed_hints

            # Build linking prompts
            code_msgs = linking.build_code_gen_prompt(
                schemas, labels, question, join_hints=all_hints,
                error_feedback=None, history=history, prev_result_meta=prev_result_meta,
            )

            # Execute with all dataframes passed dynamically
            df_kwargs = {f"df_{slot}": df for slot, df in dfs.items()}
            success, answer, chart, debug, raw_result, result_str, metadata, formatting = _run_pipeline(
                code_msgs,
                lambda r, m: linking.build_answer_gen_prompt(question, r, labels, history, m),
                debug, prev_result=prev_result, **df_kwargs
            )

    else:
        return jsonify({"error": f"Unknown mode: {mode}. This service supports 'standard' and 'variance'."}), 400

    if success:
        history.append({"role": "user", "content": question})
        history.append({"role": "assistant", "content": answer})
        sess.last_result = raw_result
        # Stored so /ask/refine can re-run Call 2 on the SAME computed result.
        sess.last_question   = question
        sess.last_result_str = result_str
        sess.last_metadata   = metadata
        sess.last_mode       = mode
        sess.last_formatting = formatting
        save_session(sid, sess)

    return jsonify({
        "answer": answer,
        "chart":  chart,
        "debug":  [{"label": d.label, "content": d.content} for d in debug],
        "formatting": formatting,
    })


@api.route("/ask/refine", methods=["POST"])
def ask_refine():
    """
    Re-explain the LAST computed result in a different style ("brief" or
    "detail"). Only Call 2 runs — the pandas result is reused unchanged, so the
    numbers cannot change. Uses the existing prompt builders untouched.
    """
    sid  = _sid()
    sess = get_session(sid)

    body  = request.get_json() or {}
    style = body.get("style", "brief")

    question   = getattr(sess, "last_question", None)
    result_str = getattr(sess, "last_result_str", None)
    metadata   = getattr(sess, "last_metadata", None)
    mode       = getattr(sess, "last_mode", "standard")

    if not question or result_str is None:
        return jsonify({"error": "Nothing to refine yet. Ask a question first."}), 400

    # Per-user key + model + usage tracking (same additive envelope as /ask)
    username = _username()
    ctx, err = _resolve_user_context(username, sess)
    if err:
        return err
    user, user_key, model_name = ctx
    request_id = uuid.uuid4().hex
    ctx_tokens = request_context.begin(user_key, model_name)
    try:
        if style == "detail":
            styled_q = (f"{question}\n\n(Re-explain the result above in MORE detail: "
                        f"cover the notable values, comparisons, and any caveats.)")
        else:
            styled_q = (f"{question}\n\n(Re-explain the result above in ONE short "
                        f"sentence — the single most important takeaway.)")

        history = sess.get_history(mode)
        if mode == "variance":
            messages = variance.build_answer_gen_prompt(styled_q, result_str, sess.label_a, sess.label_b, history, metadata)
        else:
            messages = standard.build_answer_gen_prompt(styled_q, result_str, history, metadata)

        ok, raw_answer = llm_service.generate_human_answer(messages)
        if not ok:
            return jsonify({"error": raw_answer}), 502

        answer, chart, formatting = llm_service.parse_answer_response(raw_answer)
        # Keep previously requested highlighting unless the refine answer
        # itself carries new rules — a "one sentence" rephrase must not
        # silently erase formatting the user asked for earlier.
        if formatting:
            sess.last_formatting = formatting
            save_session(sid, sess)
        effective_formatting = formatting or getattr(sess, "last_formatting", None)
        logger.info(f"Refine ({style}) | sid={sid} | mode={mode}")
        return jsonify({"answer": answer, "chart": chart, "formatting": effective_formatting})
    finally:
        usage_capture.flush(user["UserID"], username, sid, request_id, mode, question, refine=True)
        request_context.end(ctx_tokens)


@api.route("/clear", methods=["POST"])
def clear():
    sid  = _sid()
    sess = get_session(sid)
    mode = (request.get_json() or {}).get("mode")
    sess.clear_history(mode)
    save_session(sid, sess)
    logger.info(f"Chat history cleared | sid={sid} | mode={mode or 'all'}")
    return jsonify({"ok": True})


# ── Export (verbatim from prototype export.py) ────────────────────────────────

def _convert_to_dataframe(raw_result) -> pd.DataFrame:
    if isinstance(raw_result, pd.DataFrame):
        return raw_result
    if isinstance(raw_result, pd.Series):
        return raw_result.to_frame(name=raw_result.name or "Value")
    if isinstance(raw_result, dict):
        if raw_result and all(isinstance(v, (list, tuple)) for v in raw_result.values()):
            return pd.DataFrame(raw_result)
        return pd.DataFrame(list(raw_result.items()), columns=["Key", "Value"])
    if isinstance(raw_result, list) and raw_result and isinstance(raw_result[0], dict):
        return pd.DataFrame(raw_result)
    if isinstance(raw_result, list):
        return pd.DataFrame(raw_result, columns=["Value"])
    return pd.DataFrame({"Result": [raw_result]})


# Coloring loops run per cell in Python; beyond this many rows the export
# is delivered plain (instantly) instead of making the user wait minutes.
FORMATTING_MAX_ROWS = 50000


def _apply_conditional_formatting(worksheet, df):
    """
    Apply smart conditional formatting to Excel worksheet based on heuristics.

    Survey-driven feature: Users requested conditional formatting like IMS Health Check.
    This function applies automatic color coding based on column names and data patterns.
    Always enabled (no UI toggle) - users can remove formatting in Excel if not desired.

    Heuristics Applied:

    1. STATUS/HEALTH COLUMNS:
       Detects: "status", "health", "result", "grade", "check", "outcome" in column name
       Colors:
         - Green: "pass", "success", "ok", "good", "healthy", "complete", "approved"
         - Yellow: "warning", "caution", "medium", "pending", "review"
         - Red: "fail", "error", "critical", "bad", "failed", "rejected"

    2. PERCENTAGE COLUMNS:
       Detects: "%" in column name OR "percent"/"pct" in column name, AND at
       least 80% of the values inside 0-1 or 0-100. "rate" is deliberately
       excluded (a finance "rate" is usually dollars, not a percentage).
       Colors:
         - Green: >= 80%
         - Yellow: 50-79%
         - Red: < 50%

    (Score/risk/priority percentile coloring was removed — arbitrary-scale
    colors confused users. LLM-requested rules cover those cases on demand.)

    Design Notes:
    - Non-intrusive: Only formats columns that match patterns (other columns untouched)
    - Graceful: Handles missing data, mixed types, edge cases
    - Fast: Minimal overhead (<100ms for typical datasets)
    - Safe: Wrapped in try/except - formatting failure doesn't break export

    Args:
        worksheet: openpyxl worksheet object
        df: pandas DataFrame being exported
    """
    if len(df) > FORMATTING_MAX_ROWS:
        logger.info(f"Skipping auto-formatting: {len(df):,} rows exceeds the {FORMATTING_MAX_ROWS:,}-row limit")
        return

    try:
        from openpyxl.styles import PatternFill

        # Excel standard colors (subtle, professional)
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # Light green
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # Light yellow
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Light red

        # Status keywords and their mappings. Cell values must equal one of
        # these WHOLE values (case-insensitive) — substring matching used to
        # color unrelated text (e.g. "compassion" contains "pass").
        status_keywords = ["status", "health", "result", "grade", "check", "outcome"]
        pass_values = {"pass", "passed", "success", "successful", "ok", "good", "healthy",
                       "complete", "completed", "approved", "yes", "true", "on track", "green"}
        warning_values = {"warning", "caution", "medium", "pending", "in review", "review",
                          "moderate", "at risk", "yellow"}
        fail_values = {"fail", "failed", "error", "critical", "bad", "rejected", "no",
                       "false", "late", "red"}

        # Percentage keywords. "rate" is deliberately NOT here — in finance
        # data a "rate" is usually a dollar rate (labor rate), not a percentage.
        percentage_keywords = ["percent", "pct"]

        logger.info(f"Applying conditional formatting to {len(df.columns)} columns, {len(df)} rows")

        for col_idx, col_name in enumerate(df.columns):
            col_name_lower = str(col_name).lower()

            # RULE 1: Status/Health Columns
            if any(keyword in col_name_lower for keyword in status_keywords):
                logger.debug(f"Applying status formatting to column '{col_name}'")
                for row_idx in range(2, len(df) + 2):  # Skip header row (row 1)
                    cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                    value_str = str(cell.value).lower().strip() if cell.value is not None else ""

                    if value_str in pass_values:
                        cell.fill = green_fill
                    elif value_str in warning_values:
                        cell.fill = yellow_fill
                    elif value_str in fail_values:
                        cell.fill = red_fill

            # RULE 2: Percentage Columns — the name must look like a percentage
            # AND the values must actually live in 0-1 or 0-100; otherwise skip
            # (prevents coloring dollar amounts or variance columns).
            elif "%" in col_name or any(kw in col_name_lower for kw in percentage_keywords):
                numeric = pd.to_numeric(
                    df.iloc[:, col_idx].astype(str).str.replace("%", "", regex=False),
                    errors="coerce").dropna()
                if len(numeric) == 0 or ((numeric >= 0) & (numeric <= 100)).mean() < 0.8:
                    continue
                logger.debug(f"Applying percentage formatting to column '{col_name}'")
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet.cell(row=row_idx, column=col_idx + 1)
                    try:
                        # Handle both "85%" and 0.85 formats
                        value = cell.value
                        if value is None:
                            continue

                        if isinstance(value, str):
                            # Strip % sign and convert
                            value = float(value.replace("%", "").strip())
                        else:
                            value = float(value)

                        # If value is between 0-1, assume it's decimal (0.85 = 85%)
                        if 0 <= value <= 1:
                            value = value * 100

                        # Apply thresholds
                        if value >= 80:
                            cell.fill = green_fill
                        elif value >= 50:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    except (ValueError, TypeError):
                        # Skip cells that can't be converted to numbers
                        continue

            # (Former RULE 3 — percentile coloring of score/risk/priority
            # columns — was removed on purpose: coloring arbitrary thirds of a
            # numeric scale confused more than it helped. Users who want these
            # colored can ask in the question, which flows through the
            # LLM-requested formatting rules instead.)

        logger.info("Conditional formatting applied successfully")

    except Exception as e:
        # Formatting is a nice-to-have, not critical - don't break the export
        logger.warning(f"Conditional formatting failed (non-critical): {e}")
        # Continue with export - user gets plain Excel if formatting fails


def _apply_llm_formatting(worksheet, df, formatting_rules):
    """
    Apply user-requested formatting rules from LLM Call 2 response.

    Phase 2 feature (Enhanced): Supports all 3 tiers of conditional formatting:
    - Tier 1: Numeric, Text, Null/Empty, Boolean
    - Tier 2: Text contains, Row-level, Date comparisons
    - Tier 3: Top/Bottom N, Multiple conditions (AND), Cross-column

    Args:
        worksheet: openpyxl worksheet object
        df: pandas DataFrame being exported
        formatting_rules: dict with "rules" list and optional "row_level" flag

    Supported rule types:
        1. Numeric: {"column": "Sales", "condition": "<", "value": 1000000, "color": "red"}
        2. Text: {"column": "Owner", "condition": "==", "value": "John", "color": "yellow"}
        3. Null: {"column": "Email", "condition": "is_null", "color": "red"}
        4. Date: {"column": "DueDate", "condition": ">", "value": "2024-01-01", "color": "red"}
        5. Top/Bottom N: {"column": "Score", "condition": "top_n", "value": 5, "color": "green"}
        6. Cross-column: {"column": "Actual", "condition": ">", "compare_column": "Budget", "color": "red"}
        7. Multiple (AND): {"conditions": [{...}, {...}], "operator": "and", "color": "red"}
    """
    if not formatting_rules or not isinstance(formatting_rules, dict):
        return

    rules = formatting_rules.get("rules", [])
    row_level = formatting_rules.get("row_level", False)

    if not rules:
        return

    if len(df) > FORMATTING_MAX_ROWS:
        logger.info(f"Skipping requested formatting: {len(df):,} rows exceeds the {FORMATTING_MAX_ROWS:,}-row limit")
        return

    try:
        from openpyxl.styles import PatternFill

        # Same color scheme as heuristics
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        color_map = {"green": green_fill, "yellow": yellow_fill, "red": red_fill}

        logger.info(f"Applying {len(rules)} LLM formatting rules (row_level={row_level})")

        for rule in rules:
            try:
                # Handle multiple conditions (AND logic)
                if "conditions" in rule:
                    _apply_multi_condition_rule(worksheet, df, rule, color_map, row_level)
                    continue

                column = rule.get("column")
                condition = rule.get("condition")
                value = rule.get("value")
                color = rule.get("color")
                compare_column = rule.get("compare_column")  # For cross-column

                # Validate
                if not color or color not in color_map:
                    logger.warning(f"Invalid/unknown color in rule: {rule}")
                    continue

                fill = color_map[color]

                # Handle different rule types
                if condition in ["top_n", "bottom_n"]:
                    _apply_top_bottom_rule(worksheet, df, column, condition, value, fill, row_level)
                elif compare_column:
                    _apply_cross_column_rule(worksheet, df, column, condition, compare_column, fill, row_level)
                elif condition in ["is_null", "is_not_null"]:
                    _apply_null_rule(worksheet, df, column, condition, fill, row_level)
                else:
                    # Standard single-column condition
                    _apply_standard_rule(worksheet, df, column, condition, value, fill, row_level)

            except Exception as e:
                logger.warning(f"Failed to apply rule {rule}: {e}")
                continue

        logger.info("LLM formatting rules applied successfully")

    except Exception as e:
        logger.warning(f"LLM formatting failed (non-critical): {e}")


def _apply_standard_rule(worksheet, df, column, condition, value, fill, row_level):
    """Apply standard single-column condition (numeric, text, date)."""
    if column not in df.columns:
        logger.warning(f"Column '{column}' not found")
        return

    col_idx = df.columns.get_loc(column)

    for row_idx in range(2, len(df) + 2):  # Skip header
        cell = worksheet.cell(row=row_idx, column=col_idx + 1)
        cell_value = cell.value

        try:
            match = False

            # Numeric conditions
            if condition in ["<", ">", "<=", ">=", "==", "!="]:
                # Try numeric first
                try:
                    num_value = float(cell_value) if cell_value is not None else None
                    target_value = float(value) if value is not None else None
                    if num_value is not None and target_value is not None:
                        if condition == "<" and num_value < target_value:
                            match = True
                        elif condition == ">" and num_value > target_value:
                            match = True
                        elif condition == "<=" and num_value <= target_value:
                            match = True
                        elif condition == ">=" and num_value >= target_value:
                            match = True
                        elif condition == "==" and num_value == target_value:
                            match = True
                        elif condition == "!=" and num_value != target_value:
                            match = True
                except (ValueError, TypeError):
                    # Try text comparison
                    str_value = str(cell_value).strip().lower() if cell_value else ""
                    target_str = str(value).strip().lower() if value else ""
                    if condition == "==" and str_value == target_str:
                        match = True
                    elif condition == "!=" and str_value != target_str:
                        match = True
                    # Try date comparison
                    try:
                        date_value = pd.to_datetime(cell_value)
                        target_date = pd.to_datetime(value)
                        if condition == "<" and date_value < target_date:
                            match = True
                        elif condition == ">" and date_value > target_date:
                            match = True
                        elif condition == "<=" and date_value <= target_date:
                            match = True
                        elif condition == ">=" and date_value >= target_date:
                            match = True
                        elif condition == "==" and date_value == target_date:
                            match = True
                    except:
                        pass

            # Text conditions
            elif condition in ["contains", "startswith", "endswith"]:
                str_value = str(cell_value).strip().lower() if cell_value else ""
                target_str = str(value).strip().lower() if value else ""
                if condition == "contains" and target_str in str_value:
                    match = True
                elif condition == "startswith" and str_value.startswith(target_str):
                    match = True
                elif condition == "endswith" and str_value.endswith(target_str):
                    match = True

            if match:
                if row_level:
                    _highlight_row(worksheet, row_idx, len(df.columns), fill)
                else:
                    cell.fill = fill

        except Exception as e:
            continue


def _apply_null_rule(worksheet, df, column, condition, fill, row_level):
    """Apply null/empty check."""
    if column not in df.columns:
        logger.warning(f"Column '{column}' not found")
        return

    col_idx = df.columns.get_loc(column)

    for row_idx in range(2, len(df) + 2):
        cell = worksheet.cell(row=row_idx, column=col_idx + 1)
        cell_value = cell.value

        match = False
        if condition == "is_null" and (cell_value is None or str(cell_value).strip() == ""):
            match = True
        elif condition == "is_not_null" and cell_value is not None and str(cell_value).strip() != "":
            match = True

        if match:
            if row_level:
                _highlight_row(worksheet, row_idx, len(df.columns), fill)
            else:
                cell.fill = fill


def _apply_top_bottom_rule(worksheet, df, column, condition, n, fill, row_level):
    """Apply top N or bottom N highlighting."""
    if column not in df.columns:
        logger.warning(f"Column '{column}' not found")
        return

    try:
        n = int(n) if n else 5
    except:
        n = 5

    col_data = pd.to_numeric(df[column], errors='coerce')

    if condition == "top_n":
        threshold = col_data.nlargest(n).min()
        matching_rows = df[col_data >= threshold].index
    else:  # bottom_n
        threshold = col_data.nsmallest(n).max()
        matching_rows = df[col_data <= threshold].index

    col_idx = df.columns.get_loc(column)

    for idx in matching_rows:
        row_idx = idx + 2  # +2 for header and 0-indexing
        if row_level:
            _highlight_row(worksheet, row_idx, len(df.columns), fill)
        else:
            worksheet.cell(row=row_idx, column=col_idx + 1).fill = fill


def _apply_cross_column_rule(worksheet, df, column, condition, compare_column, fill, row_level):
    """Apply cross-column comparison."""
    if column not in df.columns or compare_column not in df.columns:
        logger.warning(f"Column '{column}' or '{compare_column}' not found")
        return

    col_idx = df.columns.get_loc(column)

    col1_data = pd.to_numeric(df[column], errors='coerce')
    col2_data = pd.to_numeric(df[compare_column], errors='coerce')

    for row_idx in range(2, len(df) + 2):
        df_idx = row_idx - 2
        val1 = col1_data.iloc[df_idx]
        val2 = col2_data.iloc[df_idx]

        if pd.isna(val1) or pd.isna(val2):
            continue

        match = False
        if condition == "<" and val1 < val2:
            match = True
        elif condition == ">" and val1 > val2:
            match = True
        elif condition == "<=" and val1 <= val2:
            match = True
        elif condition == ">=" and val1 >= val2:
            match = True
        elif condition == "==" and val1 == val2:
            match = True
        elif condition == "!=" and val1 != val2:
            match = True

        if match:
            if row_level:
                _highlight_row(worksheet, row_idx, len(df.columns), fill)
            else:
                worksheet.cell(row=row_idx, column=col_idx + 1).fill = fill


def _apply_multi_condition_rule(worksheet, df, rule, color_map, row_level):
    """Apply multiple conditions with AND logic."""
    conditions = rule.get("conditions", [])
    operator = rule.get("operator", "and")
    color = rule.get("color")

    if not conditions or not color or color not in color_map:
        return

    fill = color_map[color]

    # Evaluate each row
    for row_idx in range(2, len(df) + 2):
        df_idx = row_idx - 2
        all_match = True

        for cond in conditions:
            column = cond.get("column")
            condition = cond.get("condition")
            value = cond.get("value")

            if column not in df.columns:
                all_match = False
                break

            cell_value = df[column].iloc[df_idx]

            # Evaluate this condition
            match = _evaluate_single_condition(cell_value, condition, value)

            if operator == "and" and not match:
                all_match = False
                break

        if all_match:
            if row_level:
                _highlight_row(worksheet, row_idx, len(df.columns), fill)
            else:
                # Highlight first column of the multi-condition
                first_col = conditions[0].get("column")
                if first_col in df.columns:
                    col_idx = df.columns.get_loc(first_col)
                    worksheet.cell(row=row_idx, column=col_idx + 1).fill = fill


def _evaluate_single_condition(cell_value, condition, target_value):
    """Evaluate a single condition for multi-condition rules."""
    try:
        # Numeric
        if condition in ["<", ">", "<=", ">=", "==", "!="]:
            try:
                num_val = float(cell_value) if cell_value is not None else None
                target_num = float(target_value) if target_value is not None else None
                if num_val is not None and target_num is not None:
                    if condition == "<": return num_val < target_num
                    elif condition == ">": return num_val > target_num
                    elif condition == "<=": return num_val <= target_num
                    elif condition == ">=": return num_val >= target_num
                    elif condition == "==": return num_val == target_num
                    elif condition == "!=": return num_val != target_num
            except:
                # Fall back to text
                str_val = str(cell_value).strip().lower() if cell_value else ""
                target_str = str(target_value).strip().lower() if target_value else ""
                if condition == "==": return str_val == target_str
                elif condition == "!=": return str_val != target_str

        # Text
        elif condition in ["contains", "startswith", "endswith"]:
            str_val = str(cell_value).strip().lower() if cell_value else ""
            target_str = str(target_value).strip().lower() if target_value else ""
            if condition == "contains": return target_str in str_val
            elif condition == "startswith": return str_val.startswith(target_str)
            elif condition == "endswith": return str_val.endswith(target_str)

        # Null
        elif condition == "is_null":
            return cell_value is None or str(cell_value).strip() == ""
        elif condition == "is_not_null":
            return cell_value is not None and str(cell_value).strip() != ""

    except:
        pass

    return False


def _highlight_row(worksheet, row_idx, num_cols, fill):
    """Highlight entire row."""
    for col_idx in range(1, num_cols + 1):
        worksheet.cell(row=row_idx, column=col_idx).fill = fill


@api.route("/export/last_result", methods=["GET"])
def export_last_result():
    """
    Export last query result to Excel with automatic conditional formatting.

    Heuristic-based formatting is always applied (Phase 1 of survey-driven feature).
    Detects status, percentage, and score columns automatically.
    """
    sid  = _sid()
    sess = get_session(sid)
    if sess.last_result is None:
        return jsonify({"error": "No data available to export. Run a query first."}), 400
    try:
        df = _convert_to_dataframe(sess.last_result)
        if len(df) > 500000:
            return jsonify({"error": f"Result too large to export ({len(df):,} rows). Maximum is 500,000 rows."}), 400
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Query Result", index=False)
            # Apply conditional formatting (Phase 1: Heuristics)
            worksheet = writer.sheets["Query Result"]
            _apply_conditional_formatting(worksheet, df)
            # Apply LLM formatting rules (Phase 2) - overlays on heuristics
            formatting = getattr(sess, "last_formatting", None)
            if formatting:
                _apply_llm_formatting(worksheet, df, formatting)
        output.seek(0)
        logger.info(f"Excel export | sid={sid} | rows={len(df)} | cols={len(df.columns)}")
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"query_result_{sid[:8]}.xlsx")
    except Exception as e:
        logger.error(f"Export failed: {str(e)}")
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


@api.route("/export/debug_result", methods=["GET"])
def export_debug_result():
    """
    Export debug result (raw query data, possibly multi-sheet) as plain Excel.

    No conditional formatting applied - debug exports are kept plain for raw data inspection.
    """
    sid  = _sid()
    sess = get_session(sid)
    if sess.last_result is None:
        return jsonify({"error": "No data available to export. Run a query first."}), 400
    try:
        raw_result = sess.last_result
        output = io.BytesIO()
        if isinstance(raw_result, dict) and raw_result and all(isinstance(v, pd.DataFrame) for v in raw_result.values()):
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                total_rows = 0
                for sheet_name, df in raw_result.items():
                    safe_sheet_name = str(sheet_name)[:31]
                    df.to_excel(writer, sheet_name=safe_sheet_name, index=False)
                    total_rows += len(df)
                    if total_rows > 500000:
                        return jsonify({"error": f"Result too large to export ({total_rows:,} total rows). Maximum is 500,000 rows."}), 400
            output.seek(0)
            logger.info(f"Debug Excel export (multi-sheet) | sid={sid} | sheets={len(raw_result)} | total_rows={total_rows}")
            return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             as_attachment=True, download_name=f"debug_result_{sid[:8]}.xlsx")
        df = _convert_to_dataframe(raw_result)
        if len(df) > 500000:
            return jsonify({"error": f"Result too large to export ({len(df):,} rows). Maximum is 500,000 rows."}), 400
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Debug Result", index=False)
        output.seek(0)
        logger.info(f"Debug Excel export (single-sheet) | sid={sid} | rows={len(df)} | cols={len(df.columns)}")
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"debug_result_{sid[:8]}.xlsx")
    except Exception as e:
        logger.error(f"Debug export failed: {str(e)}")
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


@api.route("/export/conversation", methods=["POST"])
def export_conversation():
    """
    Export the on-screen conversation to a Word document (services/report_export.py).
    The browser sends exactly what it displayed — questions, answers, chart PNGs
    captured from the canvases, and optional debug traces — so the document is
    faithful to the chat. No LLM calls.
    """
    sid = _sid()
    payload = request.get_json() or {}
    messages = payload.get("messages") or []
    if not messages:
        return jsonify({"error": "Nothing to export yet. Ask a question first."}), 400
    try:
        docx_bytes = build_conversation_docx(payload)
        from datetime import datetime as _dt
        name = f"ChatWithData_{payload.get('mode', 'standard')}_{_dt.now().strftime('%Y%m%d_%H%M')}.docx"
        logger.info(f"Conversation export | sid={sid} | messages={len(messages)} | debug={bool(payload.get('include_debug'))}")
        return send_file(
            io.BytesIO(docx_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            as_attachment=True, download_name=name,
        )
    except Exception as e:
        logger.error(f"Conversation export failed: {str(e)}")
        return jsonify({"error": f"Export failed: {str(e)}"}), 500


# ── Per-user API keys (rollout batch) ─────────────────────────────────────────

def _validate_key_against_gateway(api_key: str):
    """Test a key with a cheap authenticated call to the gateway."""
    url = f"{Config.LITELLM_API_BASE}/v1/models"
    r = _requests.get(url, headers={"Authorization": f"Bearer {api_key}"}, timeout=8, verify=False)
    r.raise_for_status()


@api.route("/user/check_key", methods=["GET"])
def user_check_key():
    username = _username()
    exists = db.user_exists(username)
    if exists:
        db.update_user_last_login(username)
    return jsonify({"has_key": exists, "username": username})


@api.route("/user/save_key", methods=["POST"])
def user_save_key():
    username = _username()
    data = request.get_json() or {}
    api_key = (data.get("api_key") or "").strip()

    if not api_key:
        return jsonify({"error": "API key is required"}), 400
    if not api_key.startswith("sk-"):
        return jsonify({"error": "Invalid API key format (must start with 'sk-')"}), 400

    try:
        _validate_key_against_gateway(api_key)
    except Exception as e:
        logger.warning(f"API key validation failed for {username}: {e}")
        return jsonify({"error": "That key was rejected by the AI gateway. Check it and try again."}), 400

    try:
        encrypted = encrypt_api_key(api_key)
    except CryptoNotConfigured as e:
        return jsonify({"error": str(e)}), 503

    if db.user_exists(username):
        db.update_user_api_key(username, encrypted)
        logger.info(f"API key updated for {username}")
    else:
        db.create_user(username, encrypted)
        logger.info(f"New user registered: {username}")
    return jsonify({"success": True})


@api.route("/user/get_key", methods=["GET"])
def user_get_key():
    username = _username()
    user = db.get_user_by_username(username)
    if not user:
        return jsonify({"error": "No API key on file"}), 404
    try:
        masked = mask_api_key(decrypt_api_key(user["APIKey"]))
    except Exception:
        masked = "(unreadable — please re-enter)"
    return jsonify({"masked_key": masked, "created_at": user["CreatedAt"], "updated_at": user["UpdatedAt"]})


@api.route("/user/set_model", methods=["POST"])
def user_set_model():
    """
    Store user's model preference in their session (overrides admin default).
    Session-scoped — cleared on restart.
    """
    sid = _sid()
    sess = get_session(sid)
    body = request.get_json() or {}
    model = body.get("model", "").strip()

    # Single source of truth for selectable models: services/pricing.py
    from services.pricing import ALLOWED_MODEL_NAMES as ALLOWED_MODELS

    if not model:
        # Clear user preference (revert to admin default)
        sess.user_model_preference = None
        save_session(sid, sess)
        admin_default = get_current_model_config()["model_name"]
        logger.info(f"User {_username()} cleared model preference, reverting to admin default: {admin_default}")
        return jsonify({"success": True, "model": admin_default, "is_default": True})

    if model not in ALLOWED_MODELS:
        return jsonify({"error": f"Model not allowed. Choose from: {', '.join(ALLOWED_MODELS)}"}), 400

    sess.user_model_preference = model
    save_session(sid, sess)
    logger.info(f"User {_username()} selected model: {model}")
    return jsonify({"success": True, "model": model, "is_default": False})


@api.route("/user/get_model", methods=["GET"])
def user_get_model():
    """
    Return current active model (user preference or admin default) and available models.
    """
    sid = _sid()
    sess = get_session(sid)
    admin_default = get_current_model_config()["model_name"]
    current_model = sess.user_model_preference or admin_default
    is_default = sess.user_model_preference is None

    # Labels and price strings derive from services/pricing.py — one place
    # to edit models and rates.
    from services.pricing import model_options

    return jsonify({
        "current_model": current_model,
        "is_default": is_default,
        "admin_default": admin_default,
        "available_models": model_options()
    })


# ── Usage dashboards (rollout batch) ──────────────────────────────────────────

def _period_range():
    """Translate ?period=today|week|month into (start, end) DB timestamps."""
    period = request.args.get("period", "month")
    now = datetime.utcnow()
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    else:
        start = now - timedelta(days=30)
    return db.fmt_dt(start), db.fmt_dt(now)


@api.route("/user/usage/summary", methods=["GET"])
def user_usage_summary():
    start, end = _period_range()
    return jsonify(db.get_user_usage_summary(_username(), start, end))


@api.route("/user/usage/chart", methods=["GET"])
def user_usage_chart():
    start, end = _period_range()
    return jsonify({"data": db.get_user_usage_chart(_username(), start, end)})


@api.route("/user/usage/history", methods=["GET"])
def user_usage_history():
    limit = min(int(request.args.get("limit", 20)), 100)
    offset = int(request.args.get("offset", 0))
    return jsonify({"history": db.get_user_usage_history(_username(), limit, offset)})


# ── Admin (rollout batch) ─────────────────────────────────────────────────────

def _admin_guard():
    try:
        require_admin(_username())
        return None
    except PermissionError as e:
        return jsonify({"error": str(e)}), 403


@api.route("/admin/check", methods=["GET"])
def admin_check():
    return jsonify({"is_admin": is_admin(_username())})


@api.route("/admin/users/list", methods=["GET"])
def admin_users_list():
    guard = _admin_guard()
    if guard:
        return guard
    users = db.get_all_users()
    start, end = db.fmt_dt(datetime.utcnow() - timedelta(days=30)), db.fmt_dt(datetime.utcnow())
    for u in users:
        u["usage_summary"] = db.get_user_usage_summary(u["Username"], start, end)
    return jsonify({"users": users})


@api.route("/admin/usage/summary", methods=["GET"])
def admin_usage_summary():
    guard = _admin_guard()
    if guard:
        return guard
    start, end = _period_range()
    return jsonify(db.get_all_users_usage_summary(start, end))


@api.route("/admin/usage/by_user", methods=["GET"])
def admin_usage_by_user():
    guard = _admin_guard()
    if guard:
        return guard
    start, end = _period_range()
    limit = min(int(request.args.get("limit", 10)), 100)
    return jsonify({"data": db.get_usage_by_user(start, end, limit)})


@api.route("/admin/usage/by_model", methods=["GET"])
def admin_usage_by_model():
    guard = _admin_guard()
    if guard:
        return guard
    start, end = _period_range()
    return jsonify({"data": db.get_usage_by_model(start, end)})


@api.route("/admin/models/available", methods=["GET"])
def admin_models_available():
    guard = _admin_guard()
    if guard:
        return guard
    username = _username()
    user = db.get_user_by_username(username)
    try:
        api_key = decrypt_api_key(user["APIKey"]) if user else Config._ENV_LITELLM_API_KEY
    except Exception:
        api_key = Config._ENV_LITELLM_API_KEY
    return jsonify({"models": get_available_models(api_key)})


@api.route("/admin/models/current", methods=["GET"])
def admin_models_current():
    # Any signed-in user may see which model is active.
    return jsonify(get_current_model_config())


@api.route("/admin/models/select", methods=["POST"])
def admin_models_select():
    guard = _admin_guard()
    if guard:
        return guard
    data = request.get_json() or {}
    model_name = (data.get("model_name") or "").strip()
    reason = (data.get("reason") or "").strip()
    if not model_name:
        return jsonify({"error": "model_name is required"}), 400
    if change_model(model_name, _username(), reason):
        return jsonify({"success": True, "model_name": model_name})
    return jsonify({"error": "Failed to change model"}), 500


@api.route("/admin/models/history", methods=["GET"])
def admin_models_history():
    guard = _admin_guard()
    if guard:
        return guard
    return jsonify({"history": get_model_history(min(int(request.args.get("limit", 20)), 100))})


# ── Data Experts (rollout batch 2) ────────────────────────────────────────────

class _DiskFile:
    """Wraps a stored dataset file so the EXISTING upload handler can consume it."""

    def __init__(self, path: str, filename: str):
        self._path = path
        self.filename = filename

    def read(self) -> bytes:
        with open(self._path, "rb") as fh:
            return fh.read()


class _BytesFile:
    """Wraps in-memory CSV bytes (a query-backed expert's fetched result) so
    the EXISTING upload handler can consume it — same shim idea as _DiskFile."""

    def __init__(self, filename: str, data: bytes):
        self.filename = filename
        self._data = data

    def read(self) -> bytes:
        return self._data


def _prepare_source(body: dict, existing: dict = None):
    """
    Normalize an admin-submitted source definition (azure/sqlserver) into the
    stored SourceConfig JSON. SQL-login passwords are Fernet-encrypted like user
    API keys; an empty password on edit keeps the stored one (from `existing`).
    Returns (source_type, config_json, sql_query, error_response).
    """
    source_type = (body.get("source_type") or "").strip()
    if source_type not in ("azure", "sqlserver"):
        return None, None, None, (jsonify({"error": "source_type must be 'azure' or 'sqlserver'."}), 400)
    sql_query = (body.get("sql_query") or "").strip()
    cfg_in = body.get("source_config") or {}

    if source_type == "azure":
        blobs = [{"alias": (b.get("alias") or "").strip(), "url": (b.get("url") or "").strip()}
                 for b in (cfg_in.get("blobs") or [])]
        config = {"blobs": blobs}
    else:
        config = {
            "server": (cfg_in.get("server") or "").strip(),
            "database": (cfg_in.get("database") or "").strip(),
            "auth": (cfg_in.get("auth") or "windows").strip(),
        }
        if config["auth"] == "sql":
            config["username"] = (cfg_in.get("username") or "").strip()
            password = (cfg_in.get("password") or "").strip()
            if password:
                try:
                    config["encrypted_password"] = encrypt_api_key(password)
                except CryptoNotConfigured as e:
                    return None, None, None, (jsonify({"error": str(e)}), 503)
            elif existing:
                prev = data_sources.parse_source_config(existing)
                if prev.get("encrypted_password"):
                    config["encrypted_password"] = prev["encrypted_password"]
    return source_type, json.dumps(config), sql_query, None


def _run_source(source_type: str, config_json: str, sql_query: str):
    """Execute a source definition; returns (df, error_response)."""
    try:
        df = data_sources.fetch_dataframe(
            {"SourceType": source_type, "SourceConfig": config_json, "SqlQuery": sql_query})
        return df, None
    except data_sources.DataSourceError as e:
        return None, (jsonify({"error": str(e)}), 400)
    except Exception as e:
        logger.error(f"Data source query failed: {e}")
        return None, (jsonify({"error": f"The query failed unexpectedly: {str(e)[:200]}"}), 500)


def _public_source_config(expert: dict) -> dict:
    """SourceConfig for the admin UI — never ship the encrypted password out."""
    cfg = data_sources.parse_source_config(expert)
    cfg["has_password"] = bool(cfg.pop("encrypted_password", None))
    return cfg


def _validate_expert_file(file_storage, sheet_name):
    """
    Parse an admin-uploaded expert file with the same reader users' uploads use.
    Returns (df, error_response). Multi-sheet workbooks require a sheet name.
    """
    ext = os.path.splitext(file_storage.filename or "")[1].lower()
    if ext not in (".xlsx", ".xls", ".csv"):
        return None, None, (jsonify({"error": "Only .xlsx, .xls or .csv files are supported."}), 400)
    file_bytes = file_storage.read()
    if ext in (".xlsx", ".xls") and not sheet_name:
        sheets = get_excel_sheets(file_bytes)
        if len(sheets) > 1:
            return None, None, (jsonify({
                "error": f"This workbook has multiple sheets ({', '.join(sheets)}). "
                         f"Enter the sheet to use in the Sheet field and try again.",
                "sheets": sheets,
            }), 400)
    try:
        df, _enc, _warn = read_file(file_bytes, file_storage.filename, sheet_name or None)
    except FileHandlerError as e:
        return None, None, (jsonify({"error": str(e)}), 400)
    return df, file_bytes, None


@api.route("/experts/list", methods=["GET"])
def experts_list():
    """Experts the CURRENT user may use (server-side access filter)."""
    username = _username()
    groups = _groups()
    out = []
    for e in experts_svc.experts_for_user(username, groups):
        out.append({
            "id": e["ExpertID"],
            "label": e["Label"],
            "description": e["Description"] or "",
            "data_as_of": experts_svc.file_as_of(e),
            "source_type": experts_svc.source_type(e),
        })
    return jsonify({"experts": out})


@api.route("/experts/load", methods=["POST"])
def experts_load():
    """
    Load an expert into this user's Standard slot — identical downstream
    behavior to uploading the file (pipeline/insights/tracking untouched).
    """
    sid = _sid()
    sess = get_session(sid)
    username = _username()
    body = request.get_json() or {}
    expert = experts_svc.get_expert(int(body.get("expert_id", 0)))

    if not expert or not expert["IsActive"] or not experts_svc.is_loadable(expert):
        return jsonify({"error": "Data expert not found."}), 404
    if not experts_svc.user_allowed(expert, username, _groups()):
        logger.warning(f"Expert access denied | user={username} | expert={expert['Label']}")
        return jsonify({"error": "You don't have access to this data expert."}), 403

    source_type = experts_svc.source_type(expert)
    if source_type == "file":
        if not os.path.exists(expert["FilePath"]):
            return jsonify({"error": "The expert's data file is missing on the server. Contact the admin."}), 500
        sess.clear_history("standard")
        disk_file = _DiskFile(expert["FilePath"], expert["OriginalFileName"] or "expert.xlsx")
        result = _handle_upload(disk_file, "df", sess, expert["SheetName"] or None)
        data_as_of = experts_svc.file_as_of(expert)
    else:
        # Query-backed expert (azure/sqlserver): fetch live, then feed the
        # result through the SAME upload path as a CSV — pipeline untouched.
        try:
            df = data_sources.fetch_dataframe(expert)
        except data_sources.DataSourceError as e:
            logger.warning(f"Expert source fetch failed | {expert['Label']} | {e}")
            return jsonify({"error": str(e)}), 502
        except Exception as e:
            logger.error(f"Expert source fetch failed | {expert['Label']} | {e}")
            return jsonify({"error": "The expert's data source could not be read. Contact the admin."}), 502
        csv_bytes = df.to_csv(index=False).encode("utf-8-sig")
        if len(csv_bytes) > Config.MAX_FILE_BYTES:
            return jsonify({"error": f"The query result exceeds the {Config.MAX_FILE_SIZE_MB} MB "
                                     f"working limit — the admin should tighten the SQL."}), 502
        sess.clear_history("standard")
        result = _handle_upload(_BytesFile(f"{expert['Label']}.csv", csv_bytes), "df", sess, None)
        data_as_of = datetime.utcnow().strftime("%Y-%m-%d %H:%M")  # fetched just now

    if "error" in result or result.get("needs_sheet_selection"):
        return jsonify({"error": result.get("error", "The expert file could not be read. Contact the admin.")}), 500

    # Present as the expert, not as a raw file
    result["filename"] = expert["Label"]
    result["expert"] = {
        "id": expert["ExpertID"],
        "label": expert["Label"],
        "description": expert["Description"] or "",
        "data_as_of": data_as_of,
        "source_type": source_type,
    }
    custom_q = experts_svc.questions_list(expert)
    if custom_q:
        result["suggestions"] = custom_q       # owner-authored chips win
    sess.filename = expert["Label"]
    save_session(sid, sess)
    logger.info(f"Expert loaded | user={username} | expert={expert['Label']} | shape=({result['rows']},{result['cols']})")
    return jsonify(result)


@api.route("/experts/groups_in_use", methods=["GET"])
def experts_groups_in_use():
    """For the .NET layer: which AD group names should be checked on the badge."""
    return jsonify({"groups": experts_svc.groups_in_use()})


@api.route("/admin/experts/list", methods=["GET"])
def admin_experts_list():
    guard = _admin_guard()
    if guard:
        return guard
    out = []
    for e in experts_svc.all_experts():
        out.append({
            "id": e["ExpertID"], "label": e["Label"], "description": e["Description"] or "",
            "file": e["OriginalFileName"] or "", "sheet": e["SheetName"] or "",
            "rows": e["Rows"], "cols": e["Cols"],
            "access_mode": e["AccessMode"],
            "allowed_users": e["AllowedUsers"] or "", "allowed_groups": e["AllowedGroups"] or "",
            "user_count": len(experts_svc.parse_list(e["AllowedUsers"] or "")),
            "questions": e["RecommendedQuestions"] or "",
            "active": bool(e["IsActive"]),
            "updated_by": e["UpdatedBy"], "updated_at": e["UpdatedAt"],
            "data_as_of": experts_svc.file_as_of(e),
            "source_type": experts_svc.source_type(e),
            "source_config": _public_source_config(e),
            "sql_query": e.get("SqlQuery") or "",
        })
    return jsonify({"experts": out})


@api.route("/admin/experts/create", methods=["POST"])
def admin_experts_create():
    guard = _admin_guard()
    if guard:
        return guard
    username = _username()
    f = request.files.get("file")
    label = (request.form.get("label") or "").strip()
    if not label:
        return jsonify({"error": "The expert needs a name."}), 400
    if not f or not f.filename:
        return jsonify({"error": "A data file is required."}), 400

    sheet_name = (request.form.get("sheet_name") or "").strip() or None
    df, file_bytes, err = _validate_expert_file(f, sheet_name)
    if err:
        return err

    expert_id = experts_svc.create_expert(
        label=label,
        description=(request.form.get("description") or "").strip(),
        sheet_name=sheet_name,
        access_mode=(request.form.get("access_mode") or "restricted").strip(),
        allowed_users=(request.form.get("allowed_users") or "").strip(),
        allowed_groups=(request.form.get("allowed_groups") or "").strip(),
        questions=request.form.get("questions") or "",
        rows=len(df), cols=len(df.columns),
        created_by=username,
    )
    path = experts_svc.save_expert_file(expert_id, f.filename, file_bytes)
    experts_svc.set_expert_file(expert_id, path, f.filename, sheet_name, len(df), len(df.columns), username)
    logger.info(f"Expert created | '{label}' by {username} | {len(df)}x{len(df.columns)}")
    return jsonify({"success": True, "id": expert_id})


@api.route("/admin/experts/update", methods=["POST"])
def admin_experts_update():
    guard = _admin_guard()
    if guard:
        return guard
    body = request.get_json() or {}
    expert = experts_svc.get_expert(int(body.get("expert_id", 0)))
    if not expert:
        return jsonify({"error": "Expert not found"}), 404
    label = (body.get("label") or "").strip()
    if not label:
        return jsonify({"error": "The expert needs a name."}), 400

    # Query-backed experts may also update their source definition. The new
    # definition is executed once before saving, so a broken SQL/URL never
    # replaces a working one (and Rows/Cols stay honest).
    if experts_svc.source_type(expert) != "file" and ("source_config" in body or "sql_query" in body):
        body.setdefault("source_type", experts_svc.source_type(expert))
        st, cfg_json, sql_query, err = _prepare_source(body, existing=expert)
        if err:
            return err
        df, err = _run_source(st, cfg_json, sql_query)
        if err:
            return err
        experts_svc.update_expert_source(expert["ExpertID"], cfg_json, sql_query,
                                         len(df), len(df.columns), _username())

    experts_svc.update_expert(
        expert["ExpertID"], label,
        (body.get("description") or "").strip(),
        (body.get("access_mode") or "restricted").strip(),
        (body.get("allowed_users") or "").strip(),
        (body.get("allowed_groups") or "").strip(),
        body.get("questions") or "",
        _username(),
    )
    return jsonify({"success": True})


@api.route("/admin/experts/test_query", methods=["POST"])
def admin_experts_test_query():
    """
    Run a source definition WITHOUT saving anything or touching any session —
    the admin console's 'Test query' button. Returns a preview so the admin can
    confirm the data looks right before creating/saving the expert.
    """
    guard = _admin_guard()
    if guard:
        return guard
    body = request.get_json() or {}
    existing = experts_svc.get_expert(int(body.get("expert_id") or 0)) if body.get("expert_id") else None
    st, cfg_json, sql_query, err = _prepare_source(body, existing=existing)
    if err:
        return err
    df, err = _run_source(st, cfg_json, sql_query)
    if err:
        return err
    preview = df.head(5).fillna("").astype(str).replace("nan", "").to_dict(orient="records")
    return jsonify({"ok": True, "rows": len(df), "cols": len(df.columns),
                    "columns": [str(c) for c in df.columns], "preview": preview})


@api.route("/admin/experts/create_query", methods=["POST"])
def admin_experts_create_query():
    """Create a query-backed expert (azure/sqlserver). The definition is
    executed once to validate it and record Rows/Cols."""
    guard = _admin_guard()
    if guard:
        return guard
    username = _username()
    body = request.get_json() or {}
    label = (body.get("label") or "").strip()
    if not label:
        return jsonify({"error": "The expert needs a name."}), 400
    st, cfg_json, sql_query, err = _prepare_source(body)
    if err:
        return err
    df, err = _run_source(st, cfg_json, sql_query)
    if err:
        return err
    expert_id = experts_svc.create_query_expert(
        label=label,
        description=(body.get("description") or "").strip(),
        source_type_=st,
        source_config_json=cfg_json,
        sql_query=sql_query,
        access_mode=(body.get("access_mode") or "restricted").strip(),
        allowed_users=(body.get("allowed_users") or "").strip(),
        allowed_groups=(body.get("allowed_groups") or "").strip(),
        questions=body.get("questions") or "",
        rows=len(df), cols=len(df.columns),
        created_by=username,
    )
    logger.info(f"Query expert created | '{label}' ({st}) by {username} | {len(df)}x{len(df.columns)}")
    return jsonify({"success": True, "id": expert_id, "rows": len(df), "cols": len(df.columns)})


@api.route("/admin/experts/replace_file", methods=["POST"])
def admin_experts_replace_file():
    guard = _admin_guard()
    if guard:
        return guard
    username = _username()
    expert = experts_svc.get_expert(int(request.form.get("expert_id", 0)))
    if not expert:
        return jsonify({"error": "Expert not found"}), 404
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "A data file is required."}), 400
    sheet_name = (request.form.get("sheet_name") or "").strip() or None
    df, file_bytes, err = _validate_expert_file(f, sheet_name)
    if err:
        return err
    path = experts_svc.save_expert_file(expert["ExpertID"], f.filename, file_bytes)
    experts_svc.set_expert_file(expert["ExpertID"], path, f.filename, sheet_name, len(df), len(df.columns), username)
    logger.info(f"Expert file replaced | '{expert['Label']}' by {username} | {len(df)}x{len(df.columns)}")
    return jsonify({"success": True, "rows": len(df), "cols": len(df.columns)})


@api.route("/admin/experts/toggle", methods=["POST"])
def admin_experts_toggle():
    guard = _admin_guard()
    if guard:
        return guard
    body = request.get_json() or {}
    ok = experts_svc.set_expert_active(int(body.get("expert_id", 0)), bool(body.get("active")), _username())
    return jsonify({"success": ok})


# ── App factory ───────────────────────────────────────────────────────────────

def create_app() -> Flask:
    app = Flask(__name__)
    app.config["MAX_CONTENT_LENGTH"] = Config.MAX_FILE_BYTES
    app.register_blueprint(api)

    # Rollout batch: usage DB + token-capture shim (llm.py itself is untouched).
    try:
        db.initialize_database()
    except Exception as e:
        logger.error(f"Database initialization failed (key/usage features disabled): {e}")
    usage_capture.install()

    @app.route("/health")
    @app.route("/api/health")
    def health():
        return jsonify({
            "ok": True,
            "service": "chat-with-data-python",
            "modes": ["standard", "variance"],
            "llm_gateway_configured": bool(Config.LITELLM_API_KEY and Config.LITELLM_API_BASE),
        })

    @app.errorhandler(MissingSessionId)
    def _missing_sid(_e):
        return jsonify({"error": "Missing X-Session-Id header (provided by the front end)."}), 400

    @app.errorhandler(413)
    def _too_large(_e):
        return jsonify({"error": f"File too large. Max {Config.MAX_FILE_SIZE_MB} MB."}), 413

    return app


app = create_app()

if __name__ == "__main__":
    port = int(__import__("os").environ.get("SERVICE_PORT", "8000"))
    app.run(host="0.0.0.0", port=port, debug=Config.DEBUG)
